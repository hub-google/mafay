import os
import glob
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def format_excel(excel_path):
    wb = load_workbook(excel_path)
    for ws in wb.worksheets:
        # Freeze top row
        ws.freeze_panes = 'A2'
        # Auto-fit columns
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[col_letter].width = adjusted_width
    wb.save(excel_path)

def main():
    print("Starting merge process...")
    master_files = glob.glob("chunk_*_master.csv")
    circle_files = glob.glob("chunk_*_circle.csv")
    bandit_files = glob.glob("chunk_*_bandit.csv")

    master_df = pd.concat([pd.read_csv(f) for f in master_files]) if master_files else pd.DataFrame()
    circle_df = pd.concat([pd.read_csv(f) for f in circle_files]) if circle_files else pd.DataFrame()
    bandit_df = pd.concat([pd.read_csv(f) for f in bandit_files]) if bandit_files else pd.DataFrame()

    excel_path = "analysis_results.xlsx"
    
    # If old results exist, merge them
    if os.path.exists(excel_path):
        print(f"Found existing {excel_path}, merging data...")
        try:
            old_master = pd.read_excel(excel_path, sheet_name="Match_Master")
            old_circle = pd.read_excel(excel_path, sheet_name="Circle_Stats")
            old_bandit = pd.read_excel(excel_path, sheet_name="Bandit_Events")
            
            master_df = pd.concat([old_master, master_df])
            circle_df = pd.concat([old_circle, circle_df])
            bandit_df = pd.concat([old_bandit, bandit_df])
        except Exception as e:
            print(f"Error reading old excel: {e}")

    # Deduplicate and Sort
    if not master_df.empty:
        master_df.drop_duplicates(subset=['Match_ID'], keep='last', inplace=True)
        if 'Game_Start_Time' in master_df.columns:
            master_df.sort_values(by=['Video_URL', 'Game_Start_Time'], inplace=True)
            
    if not circle_df.empty:
        circle_df.drop_duplicates(subset=['Match_ID', 'Circle_Level'], keep='last', inplace=True)
        
    if not bandit_df.empty:
        bandit_df.drop_duplicates(subset=['Match_ID', 'Event_Time'], keep='last', inplace=True)

    # Write to Excel
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        master_df.to_excel(writer, sheet_name="Match_Master", index=False)
        circle_df.to_excel(writer, sheet_name="Circle_Stats", index=False)
        bandit_df.to_excel(writer, sheet_name="Bandit_Events", index=False)

    # Format
    format_excel(excel_path)
    print("Merge and formatting complete.")

if __name__ == "__main__":
    main()
